from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

main=Tk()
main.title("Login Form")
main.geometry("800x500")
main.config(highlightbackground="black",highlightthickness=2)
        
file = pathlib.Path("TEST_Austritt.xlsx")
if file.exists ():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Austritt"
    sheet["B1"]="Wechsel intern"
    sheet["C1"]="Namensänderung"
    sheet["D1"]="UserIn"
    sheet["E1"]="Austrittsdatum"
    sheet["F1"]="Niederlassung"
    sheet["G1"]="AD Gruppen entfernen"
    sheet["H1"]="Vorgesetzte OneDrive"
    sheet["I1"]="Hardware"
    sheet["J1"]="Smartphone"
    sheet["K1"]="Tablet"
    sheet["L1"]="MDM entfernen"
    sheet["M1"]="Gelöscht ( nach 1 Monat)"
    sheet["N1"]="Notiz"
    sheet["O1"]="Mail an HR, bei Namensänderung"
    file.save("TEST_Austritt.xlsx")



def submit():
    austrittGet=austritt.get()
    wechselGet=wechselI.get()
    namensaenderungGet=namensaenderung.get()
    userInGet=userIn.get()
    austrittsdatumGet=austrittsdatum.get()
    niederlassungGet=niederlassung.get()
    adGet=ad.get()
    oneDriveGet=oneDrive.get()
    hardwareGet=hardware.get()
    smartphoneGet=smartphone.get()
    tabletGet=tablet.get()
    mdmGet=mdm.get()
    geloeschtGet=geloescht.get()
    notizGet=notiz.get()
    

   
    # niederlassungGet=wechselI.get()
    # wechselGet=wechselI.get()
    # z1=passEntry.get()
    # y1=emailentry.get()
    
    # zum Testen  CMD
    print(austrittGet)  
    print(wechselGet) 
    print(namensaenderungGet)
    print(userInGet)
    print(austrittsdatumGet)
    print(niederlassungGet)
    
    
    file=openpyxl.load_workbook("TEST_Austritt.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=austrittGet)
    sheet.cell(column=2,row=sheet.max_row,value=wechselGet)
    sheet.cell(column=3,row=sheet.max_row,value=namensaenderungGet)
    sheet.cell(column=4,row=sheet.max_row,value=userInGet)
    sheet.cell(column=5,row=sheet.max_row,value=austrittsdatumGet)
    sheet.cell(column=6,row=sheet.max_row,value=niederlassungGet)
    sheet.cell(column=7,row=sheet.max_row,value=adGet)
    sheet.cell(column=8,row=sheet.max_row,value=oneDriveGet)
    sheet.cell(column=9,row=sheet.max_row,value=hardwareGet)
    sheet.cell(column=10,row=sheet.max_row,value=smartphoneGet)
    sheet.cell(column=11,row=sheet.max_row,value=tabletGet)
    sheet.cell(column=12,row=sheet.max_row,value=mdmGet)
    sheet.cell(column=13,row=sheet.max_row,value=geloeschtGet)
    sheet.cell(column=14,row=sheet.max_row,value=notizGet)
    sheet.cell(column=15,row=sheet.max_row,value=oneDriveGet)

    # sheet.cell(column=3,row=sheet.max_row,value=z1)
    # sheet.cell(column=4,row=sheet.max_row,value=y1)
    
    # IF Else Abfrage evtl. für austritt usw. verwenden
    # if var1.get()==1:
    #     gen="Male"
    #     print("Male")
    #     sheet.cell(column=5,row=sheet.max_row,value="Male")
    # else:
    #     print("female")
    #     sheet.cell(column=5,row=sheet.max_row,value="Female")
    
    # if var2.get() ==1:
    #     print("Standard")
    #     sheet.cell(column=6,row=sheet.max_row,value="Standard") 
    # if var3.get() ==1:
    #     print("Premium")
    #     sheet.cell(column=6,row=sheet.max_row,value="Premium")         
    file.save("TEST_Austritt.xlsx")
'''
    xlfile = pd.read_excel('Backened_Data.xlsx', 'Sheet') # reading xl file 
    xlfile.to_csv('Backened_Data.csv', index=False)#conversion to csv
'''
    
  
frame1 = LabelFrame(main, text = 'Austrittsliste:').pack(expand = 'yes', fill = 'both')
Label(frame1,text="Austritt:").place(x=50,y=30)
Label(frame1,text="Wechsel Intern:").place(x=50,y=60)
Label(frame1,text="Namensänderung:").place(x=50,y=90)
Label(frame1,text="UserIn:").place(x=50,y=120)
Label(frame1,text="Austrittsdatum:").place(x=50,y=150)
Label(frame1,text="Niederlassung:").place(x=50,y=180)
Label(frame1,text="AD Gruppen:").place(x=50,y=210)
Label(frame1,text="OneDrive:").place(x=50,y=240)
Label(frame1,text="Hardware:").place(x=50,y=270)
Label(frame1,text="Smartphone:").place(x=50,y=300)
Label(frame1,text="Tablet:").place(x=50,y=330)
Label(frame1,text="MDM:").place(x=50,y=360)
Label(frame1,text="Geloescht:").place(x=50,y=390)
Label(frame1,text="Notiz:").place(x=50,y=420)
# Label(frame1,text="Password:").place(x=50,y=110)
# Label(main,text="Mail ID:").place(x=50,y=150)

#Definition def submit
austritt = Entry(frame1)
austritt.place(x=250,y=30)
wechselI =Entry(frame1)
wechselI.place(x=250,y=60)
namensaenderung =Entry(frame1)
namensaenderung.place(x=250,y=90)
userIn =Entry(frame1)
userIn.place(x=250,y=120)
austrittsdatum =Entry(frame1)
austrittsdatum.place(x=250,y=150)
niederlassung =Entry(frame1)
niederlassung.place(x=250,y=180)
ad =Entry(frame1)
ad.place(x=250,y=210)
oneDrive =Entry(frame1)
oneDrive.place(x=250,y=240)
hardware =Entry(frame1)
hardware.place(x=250,y=270)
smartphone =Entry(frame1)
smartphone.place(x=250,y=300)
tablet =Entry(frame1)
tablet.place(x=250,y=330)
mdm =Entry(frame1)
mdm.place(x=250,y=360)
geloescht =Entry(frame1)
geloescht.place(x=250,y=390)
notiz =Entry(frame1)
notiz.place(x=250,y=420)


# password = StringVar() 
# passEntry = Entry(frame1, textvariable=password, show='*')
# passEntry.place(x=250,y=110)
# emailentry = Entry(frame1)
# emailentry.place(x=250,y=150,width=250)
# var =IntVar()
# var.set('0')


#Bereich unten 2tes Frame
# frame2 = LabelFrame(main, text = 'Other Details:').pack(expand = 'yes', fill = 'both')

# label_3 = Label(frame2, text="Gender",width=20,font=("bold", 10))
# label_3.place(x=50,y=300)
# var1 = IntVar()
# Radiobutton(frame2, text="Male",padx = 5, variable=var1, value=1).place(x=200,y=300)
# Radiobutton(frame2, text="Female",padx = 20, variable=var1, value=2).place(x=280,y=300)
# label_4 = Label(frame2, text="Subscription",width=20,font=("bold", 10))
# label_4.place(x=50,y=360)
# var2 = IntVar()
# Checkbutton(frame2, text="Standard", variable=var2).place(x=200,y=360)
# var3 = IntVar()
# Checkbutton(frame2, text="Premium", variable=var3).place(x=300,y=360)
# Button(frame2,text = "Subscribe",command = submit).place(x=400,y=400)

# var1 = IntVar()
# var2 = IntVar()
# var3 = IntVar()
#var4 = IntVar()
Button(frame1,text = "Subscribe",command = submit).place(x=400,y=400)

main.mainloop()
